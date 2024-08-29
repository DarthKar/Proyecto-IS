from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from control_bd import BaseDatos
from datetime import datetime
from PySide6.QtWidgets import QMessageBox
from PySide6.QtCore import QStandardPaths

class exportarControl():
   def __init__(self) -> None:
      pass

   def exportarInventario(self):
      try:
         libro = Workbook()
         pag = libro.active
         datos = BaseDatos.leerFilas()
         horaActual = datetime.now()
         fecha = horaActual.strftime("%Y-%m-%d")
         hora = horaActual.strftime("%H:%M")

         pag["A1"] = "Nombre"
         pag["C1"] = "Cantidad"
         pag["D1"] = "Und/Medida"

         for i in range(2, len(datos) + 2):
            pag.merge_cells(f"A{i}:B{i}")
            pag[f"A{i}"] = datos[i - 2][0]
            pag[f"C{i}"] = datos[i - 2][1]
            pag[f"D{i}"] = datos[i - 2][2]

         pag.merge_cells(f"A{len(datos) + 2}:C{len(datos) + 2}")
         pag[f"A{len(datos) + 2}"] = f"Hora de generacion: {hora}"

         registros = BaseDatos.leerRegistros()
         pag.merge_cells("F1:G1")
         pag["F1"] = "Registro de cambios"

         for j in range(2, len(registros) + 2):
            pag[f"F{j}"] = registros[j - 2][0]
            pag[f"G{j}"] = registros[j - 2][1]
            pag[f"H{j}"] = registros[j - 2][2]
            pag.merge_cells(start_row=j, start_column=8, end_row=j, end_column=15)
            pag[f"H{j}"].alignment = Alignment(wrap_text=True)

         escritorio_path = QStandardPaths.writableLocation(QStandardPaths.DesktopLocation)
         filename = f"Inventario {fecha}.xlsx"
         filepath = f"{escritorio_path}/{filename}"
         libro.save(filepath)
         QMessageBox.information(None, "Informacion", "Se genero con exito el archivo")
      except Exception as e:
         QMessageBox.warning(None, "Error", f"Ocurrio un error en la generacion del archivo: {str(e)}")

   def exportarVentas(self):
      try:
         libro = Workbook()
         pag = libro.active
         datos = BaseDatos.leerOrdenes()
         horaActual = datetime.now()
         fecha = horaActual.strftime("%Y-%m-%d")
         hora = horaActual.strftime("%H:%M")

         pag["A1"] = "N° Orden"
         pag["B1"] = "Hora/Creacion"
         pag.merge_cells("B1:C1")
         pag["D1"] = "Hora/Validacion"
         pag.merge_cells("D1:E1")
         pag["F1"] = "Valor"
         pag.merge_cells("F1:G1")
         pag["H1"] = "Informacion de la orden"
         pag.merge_cells("H1:M1")

         for i in range(2, len(datos) + 2):
            pag.merge_cells(f"B{i}:C{i}")
            pag.merge_cells(f"D{i}:E{i}")
            pag.merge_cells(f"F{i}:G{i}")
            pag[f"A{i}"] = datos[i - 2][0]
            pag[f"A{i}"].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            pag[f"B{i}"] = datos[i - 2][1]
            pag[f"B{i}"].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            pag[f"D{i}"] = datos[i - 2][2]
            pag[f"D{i}"].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left') 
            pag[f"F{i}"] = datos[i - 2][3]
            pag[f"F{i}"].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            datosOrden = BaseDatos.datosOrden(str(datos[i - 2][0]))
            orde=""
            pag.merge_cells(f"H{i}:M{i}")
            for j in range(2, len(datosOrden)+2):
               orde+=datosOrden[j-2][1]
               if j!=len(datosOrden)+1:
                  orde+=","
            pag[f"H{i}"].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
            pag[f"H{i}"] = orde
         
            max_length = max(len(line) for line in orde.split(", "))
            row_height = 15 + (max_length // 40) * 15  
            pag.row_dimensions[i].height = row_height
            
         pag.merge_cells(f"A{len(datos) + 2}:C{len(datos) + 2}")
         pag[f"A{len(datos) + 2}"] = f"Hora de generacion: {hora}"
         escritorio_path = QStandardPaths.writableLocation(QStandardPaths.DesktopLocation)
         filename = f"Ventas-{fecha}.xlsx"
         filepath = f"{escritorio_path}/{filename}"
         libro.save(filepath)
         QMessageBox.information(None, "Informacion", "Se genero con exito el archivo")
      except Exception as e:
         QMessageBox.warning(None, "Error", f"Ocurrio un error en la generacion del archivo: {str(e)}")


