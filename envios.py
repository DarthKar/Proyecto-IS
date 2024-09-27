import os
from dotenv import load_dotenv
from email.mime.base import MIMEBase
from openpyxl.styles import Font, Alignment
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText  
from email import encoders
from datetime import datetime
from io import BytesIO
import ssl
import smtplib
from openpyxl import Workbook
from control_bd import BaseDatos  

def generar_excel_inventario():
    libro = Workbook()
    pag = libro.active
    datos = BaseDatos.leerFilas() 
    horaActual = datetime.now()
    fecha = horaActual.strftime("%Y-%m-%d")
    hora = horaActual.strftime("%H:%M")
    
    pag["A1"] = "Nombre"
    pag["C1"] = "Cantidad"
    pag["D1"] = "Und/Medida"

    for i in range(2, len(datos) + 2):
        pag.merge_cells(f"A{i}:B{i}")
        pag[f"A{i}"] = datos[i - 2][0]
        pag[f"C{i}"] = datos[i - 2][1]
        pag[f"D{i}"] = datos[i - 2][2]

    pag.merge_cells(f"A{len(datos) + 2}:C{len(datos) + 2}")
    pag[f"A{len(datos) + 2}"] = f"Hora de generacion: {hora}"
    registros = BaseDatos.leerRegistros()
    pag.merge_cells("F1:G1")
    pag["F1"] = "Registro de cambios"

    for j in range(2, len(registros) + 2):
            pag[f"F{j}"] = registros[j - 2][0]
            pag[f"G{j}"] = registros[j - 2][1]
            pag[f"H{j}"] = registros[j - 2][2]
            pag.merge_cells(start_row=j, start_column=8, end_row=j, end_column=15)
            pag[f"H{j}"].alignment = Alignment(wrap_text=True)

    
        
    excel_en_memoria = BytesIO()
    libro.save(excel_en_memoria)
    excel_en_memoria.seek(0)
    
    return excel_en_memoria

def generar_excel_ventas():
    libro = Workbook()
    pag = libro.active
    datos = BaseDatos.leerOrdenes()
    horaActual = datetime.now()
    fecha = horaActual.strftime("%Y-%m-%d")
    hora = horaActual.strftime("%H:%M")
    pag["A1"] = "N° Orden"
    pag["B1"] = "Hora/Creacion"
    pag.merge_cells("B1:C1")
    pag["D1"] = "Hora/Validacion"
    pag.merge_cells("D1:E1")
    pag["F1"] = "Valor"
    pag.merge_cells("F1:G1")
    pag["H1"] = "Informacion de la orden"
    pag.merge_cells("H1:M1")
    for i in range(2, len(datos) + 2):
       pag.merge_cells(f"B{i}:C{i}")
       pag.merge_cells(f"D{i}:E{i}")
       pag.merge_cells(f"F{i}:G{i}")
       pag[f"A{i}"] = datos[i - 2][0]
       pag[f"A{i}"].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
       pag[f"B{i}"] = datos[i - 2][1]
       pag[f"B{i}"].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
       pag[f"D{i}"] = datos[i - 2][2]
       pag[f"D{i}"].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left') 
       pag[f"F{i}"] = datos[i - 2][3]
       pag[f"F{i}"].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
       datosOrden = BaseDatos.datosOrden(str(datos[i - 2][0]))
       orde=""
       pag.merge_cells(f"H{i}:M{i}")
       for j in range(2, len(datosOrden)+2):
          orde+=datosOrden[j-2][1]
          if j!=len(datosOrden)+1:
             orde+=","
       pag[f"H{i}"].alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
       pag[f"H{i}"] = orde
    
       max_length = max(len(line) for line in orde.split(", "))
       row_height = 15 + (max_length // 40) * 15  
       pag.row_dimensions[i].height = row_height
       
    pag.merge_cells(f"A{len(datos) + 2}:C{len(datos) + 2}")
    pag[f"A{len(datos) + 2}"] = f"Hora de generacion: {hora}"

      
    excel_en_ventas = BytesIO()
    libro.save(excel_en_ventas)
    excel_en_ventas.seek(0)

    return excel_en_ventas

def enviar_email_con_adjunto():
    try:
        load_dotenv()
        email_envi = "c1677672@gmail.com"
        email_rec = "hidekonz33@gmail.com"
        contra = os.getenv("CODIGO")
        horaActual = datetime.now()
        fecha = horaActual.strftime("%Y-%m-%d")
        hora = horaActual.strftime("%H:%M")
        sub = f"exportaciones {fecha}"
        cuerpo = f"generado {hora}"

        archivo_excel = generar_excel_inventario()
        archivo_excel2 = generar_excel_ventas()

        em = MIMEMultipart()
        em["From"] = email_envi
        em["To"] = email_rec
        em["Subject"] = sub
        em.attach(MIMEText(cuerpo, "plain"))  

        adjunto = MIMEBase("application", "octet-stream")
        adjunto.set_payload(archivo_excel.read())
        encoders.encode_base64(adjunto)
        adjunto.add_header(
            "Content-Disposition",
            f"attachment; filename=Inventario-{fecha}.xlsx"
        )

        adjunto2 = MIMEBase("application", "octet-stream")
        adjunto2.set_payload(archivo_excel2.read())
        encoders.encode_base64(adjunto2)
        adjunto2.add_header(
            "Content-Disposition",
            f"attachment; filename=Ventas-{fecha}.xlsx"
        )

        em.attach(adjunto)
        em.attach(adjunto2)

        context = ssl.create_default_context()

        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as smtp:
            smtp.login(email_envi, contra)
            smtp.sendmail(email_envi, email_rec, em.as_string())

        print("Correo enviado con éxito.")

    except Exception as e:
        print(f"Error al enviar el correo: {e}")


